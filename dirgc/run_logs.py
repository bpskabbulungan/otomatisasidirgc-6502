import os
import re
from datetime import datetime
from pathlib import Path


LOGS_DIR = "logs"
DEFAULT_LOG_TYPE = "run"
COLUMNS = [
    "no",
    "idsbr",
    "nama_usaha",
    "alamat",
    "keberadaanusaha_gc",
    "latitude",
    "latitude_source",
    "latitude_before",
    "latitude_after",
    "longitude",
    "longitude_source",
    "longitude_before",
    "longitude_after",
    "hasil_gc_before",
    "hasil_gc_after",
    "nama_usaha_before",
    "nama_usaha_after",
    "alamat_before",
    "alamat_after",
    "status",
    "catatan",
]


def _next_run_number(date_dir, prefix):
    max_run = 0
    pattern = f"{prefix}*_*.xlsx"
    for path in date_dir.glob(pattern):
        match = re.match(rf"{re.escape(prefix)}(\d+)_", path.stem)
        if not match:
            continue
        try:
            number = int(match.group(1))
        except ValueError:
            continue
        if number > max_run:
            max_run = number
    return max_run + 1


def build_run_log_path(*, now=None, log_type=None):
    now = now or datetime.now()
    log_type = (log_type or DEFAULT_LOG_TYPE).strip().lower()
    if log_type not in {"run", "update", "validasi"}:
        log_type = DEFAULT_LOG_TYPE
    date_folder = now.strftime("%Y%m%d")
    date_dir = Path(LOGS_DIR) / log_type / date_folder
    date_dir.mkdir(parents=True, exist_ok=True)
    run_number = _next_run_number(date_dir, log_type)
    time_label = now.strftime("%H%M")
    filename = f"{log_type}{run_number}_{time_label}.xlsx"
    return date_dir / filename


def write_run_log(rows, output_path):
    try:
        import pandas as pd
    except ImportError:
        pd = None

    if pd:
        df = pd.DataFrame(rows, columns=COLUMNS)
        df.to_excel(output_path, index=False)
        return

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Install pandas or openpyxl to write log Excel."
        ) from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in COLUMNS])
    workbook.save(output_path)
    workbook.close()


def append_run_log_rows(rows, output_path):
    if not rows:
        return
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Install openpyxl to append log Excel."
        ) from exc

    if not output_path:
        raise ValueError("output_path is required.")

    output_path = str(output_path)
    dir_name = os.path.dirname(output_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)

    if os.path.exists(output_path):
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
            sheet.append(COLUMNS)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(COLUMNS)

    for row in rows:
        sheet.append([row.get(col, "") for col in COLUMNS])
    workbook.save(output_path)
    workbook.close()
