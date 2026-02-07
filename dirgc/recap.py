import json
import os
import re
import shutil
import tempfile
import time

from .logging_utils import log_error, log_info, log_warn
from .settings import (
    DEFAULT_RECAP_LENGTH,
    RECAP_LENGTH_MAX,
    RECAP_LENGTH_WARN_THRESHOLD,
)


DEFAULT_OUTPUT_DIR = os.path.join("logs", "recap")
CHECKPOINT_FILENAME = "rekap_checkpoint.json"
SPLIT_SHEETS = ("Sudah GC", "Belum GC", "Duplikat")
BACKUP_SUFFIX = ".bak"


def _now_readable():
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _format_duration(seconds):
    seconds = max(0, int(seconds))
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def _format_eta_id(seconds):
    seconds = max(0, int(seconds))
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hours} Jam {minutes} Menit {secs} Detik"


def _to_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _extract_csrf_token(page):
    try:
        locator = page.locator('meta[name="csrf-token"]')
        if locator.count() > 0:
            return locator.first.get_attribute("content") or ""
    except Exception:
        return ""
    return ""


def _next_run_number(date_dir, prefix):
    max_run = 0
    pattern = f"{prefix}*_*.xlsx"
    for filename in os.listdir(date_dir):
        if not filename.endswith(".xlsx"):
            continue
        if not filename.startswith(prefix):
            continue
        stem = filename[:-5]
        match = re.match(rf"{re.escape(prefix)}(\d+)_", stem)
        if not match:
            continue
        try:
            number = int(match.group(1))
        except ValueError:
            continue
        if number > max_run:
            max_run = number
    return max_run + 1


def _build_output_path(output_dir):
    date_folder = time.strftime("%Y%m%d")
    date_dir = os.path.join(output_dir, date_folder)
    os.makedirs(date_dir, exist_ok=True)
    run_number = _next_run_number(date_dir, "rekap")
    time_label = time.strftime("%H%M%S")
    filename = f"rekap{run_number}_{time_label}.xlsx"
    return os.path.join(date_dir, filename)


def _collect_filter_payload(page, status_filter):
    keys = [
        "nama_usaha",
        "alamat_usaha",
        "provinsi",
        "kabupaten",
        "kecamatan",
        "desa",
        "status_filter",
        "rtotal",
        "sumber_data",
        "skala_usaha",
        "idsbr",
        "history_profiling",
        "f_latlong",
        "f_gc",
    ]
    payload = {}

    def pick_value(name):
        selectors = [
            f"[name='{name}']",
            f"#{name}",
            f"#f_{name}",
        ]
        for selector in selectors:
            try:
                locator = page.locator(selector)
                if locator.count() > 0:
                    return locator.first.input_value()
            except Exception:
                continue
        return ""

    for key in keys:
        if key == "status_filter":
            payload[key] = status_filter
            continue
        payload[key] = pick_value(key)

    if not payload.get("rtotal"):
        payload["rtotal"] = "0"
    return payload


def _classify_status(gcs_result):
    if gcs_result is None:
        return "Belum GC"
    value = str(gcs_result).strip()
    if not value:
        return "Belum GC"
    if value == "4":
        return "Duplikat"
    return "Sudah GC"


def _normalize_record(record, batch_start, captured_at):
    def get(key):
        return record.get(key) if isinstance(record, dict) else None

    gcs_result = get("gcs_result")
    return {
        "idsbr": get("idsbr"),
        "nama_usaha": get("nama_usaha"),
        "alamat_usaha": get("alamat_usaha"),
        "kegiatan_usaha": get("kegiatan_usaha"),
        "skala_usaha": get("skala_usaha"),
        "sumber_data": get("sumber_data"),
        "kode_wilayah": get("kode_wilayah"),
        "kdprov": get("kdprov"),
        "kdkab": get("kdkab"),
        "kdkec": get("kdkec"),
        "kddesa": get("kddesa"),
        "nmprov": get("nmprov"),
        "nmkab": get("nmkab"),
        "nmkec": get("nmkec"),
        "nmdesa": get("nmdesa"),
        "latitude": get("latitude"),
        "longitude": get("longitude"),
        "latlong_status": get("latlong_status"),
        "gcs_result": gcs_result,
        "status_gc": _classify_status(gcs_result),
        "status_perusahaan": get("status_perusahaan"),
        "gc_username": get("gc_username"),
        "latitude_gc": get("latitude_gc"),
        "longitude_gc": get("longitude_gc"),
        "latlong_status_gc": get("latlong_status_gc"),
        "history_ref_profiling_id": get("history_ref_profiling_id"),
        "perusahaan_id": get("perusahaan_id"),
        "captured_at": captured_at,
        "batch_start": batch_start,
    }


def _ensure_sheet(workbook, name, columns):
    if name in workbook.sheetnames:
        sheet = workbook[name]
    else:
        sheet = workbook.create_sheet(title=name)
    if (
        sheet.max_row == 1
        and sheet.max_column == 1
        and sheet.cell(1, 1).value is None
    ):
        sheet.append(columns)
    return sheet


def _backup_path(xlsx_path):
    return f"{xlsx_path}{BACKUP_SUFFIX}"


def _backup_existing(xlsx_path):
    if not os.path.exists(xlsx_path):
        return None
    backup_path = _backup_path(xlsx_path)
    try:
        shutil.copy2(xlsx_path, backup_path)
        return backup_path
    except Exception as exc:
        log_warn(
            "Failed to create Excel backup.",
            path=backup_path,
            error=str(exc),
        )
        return None


def _safe_save_workbook(workbook, xlsx_path):
    dir_name = os.path.dirname(xlsx_path) or "."
    base = os.path.basename(xlsx_path)
    tmp_path = None
    try:
        os.makedirs(dir_name, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(
            prefix=f".{base}.", suffix=".tmp", dir=dir_name
        )
        os.close(fd)
        workbook.save(tmp_path)
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    try:
        os.replace(tmp_path, xlsx_path)
        return
    except Exception as exc:
        fallback_path = f"{xlsx_path}.new"
        try:
            os.replace(tmp_path, fallback_path)
            log_warn(
                "Failed to replace Excel; wrote fallback file.",
                path=fallback_path,
                error=str(exc),
            )
            return
        except Exception:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        raise


def _append_xlsx_split(xlsx_path, rows, columns, *, backup_existing=True):
    if not rows:
        return
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Install openpyxl to write rekap Excel."
        ) from exc

    new_book = not os.path.exists(xlsx_path)
    if new_book:
        workbook = openpyxl.Workbook()
        active = workbook.active
        active.title = SPLIT_SHEETS[0]
    else:
        try:
            workbook = openpyxl.load_workbook(xlsx_path)
        except Exception as exc:
            backup_path = _backup_path(xlsx_path)
            if os.path.exists(backup_path):
                log_warn(
                    "Failed to load Excel; using backup.",
                    path=backup_path,
                    error=str(exc),
                )
                workbook = openpyxl.load_workbook(backup_path)
            else:
                raise

    for name in SPLIT_SHEETS:
        _ensure_sheet(workbook, name, columns)

    grouped = {name: [] for name in SPLIT_SHEETS}
    for row in rows:
        status = row.get("status_gc") or "Belum GC"
        if status not in grouped:
            status = "Belum GC"
        grouped[status].append(row)

    for name, items in grouped.items():
        if not items:
            continue
        sheet = workbook[name]
        for row in items:
            sheet.append([row.get(col, "") for col in columns])

    if backup_existing:
        _backup_existing(xlsx_path)
    _safe_save_workbook(workbook, xlsx_path)


def _load_checkpoint(path):
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return None


def _save_checkpoint(path, data):
    try:
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(data, handle, ensure_ascii=False, indent=2)
    except Exception:
        return False
    return True


def _fetch_page(page, url, payload, headers, timeout_ms=60000):
    response = page.request.post(
        url, form=payload, headers=headers, timeout=timeout_ms
    )
    status = response.status
    if status != 200:
        return status, None
    try:
        return status, response.json()
    except Exception:
        return status, None


def run_recap(
    page,
    monitor,
    *,
    status_filter="semua",
    length=DEFAULT_RECAP_LENGTH,
    output_dir=None,
    sleep_ms=800,
    max_retries=3,
    resume=True,
    backup_every=10,
    progress_callback=None,
):
    start_ts_epoch = time.time()

    def _normalize_length(value, fallback):
        try:
            value = int(value)
        except (TypeError, ValueError):
            return fallback
        return value if value > 0 else fallback

    length = _normalize_length(length, DEFAULT_RECAP_LENGTH)
    try:
        backup_every = int(backup_every)
    except (TypeError, ValueError):
        backup_every = 1
    if backup_every < 0:
        backup_every = 0
    output_dir = output_dir or DEFAULT_OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)
    checkpoint_path = os.path.join(output_dir, CHECKPOINT_FILENAME)
    legacy_checkpoint_path = os.path.join(output_dir, "recap_checkpoint.json")

    checkpoint = None
    if resume:
        checkpoint = _load_checkpoint(checkpoint_path)
        if checkpoint is None and os.path.exists(legacy_checkpoint_path):
            checkpoint = _load_checkpoint(legacy_checkpoint_path)
            if checkpoint is not None:
                log_warn(
                    "Legacy checkpoint found; resume using it once.",
                    path=legacy_checkpoint_path,
                )
    start_at = 0
    total_records = None
    output_path = None
    columns = None
    effective_records_hint = None

    if checkpoint and checkpoint.get("status") == "running":
        if checkpoint.get("status_filter") == status_filter:
            start_at = checkpoint.get("last_start", 0)
            length = checkpoint.get("length", length)
            total_records = checkpoint.get("records_total")
            effective_records_hint = checkpoint.get("records_filtered")
            output_path = checkpoint.get("output_path")
            columns = checkpoint.get("columns")
            if output_path and os.path.exists(output_path):
                log_info(
                    "Resuming rekap from checkpoint.",
                    start_at=start_at,
                    length=length,
                    output_path=output_path or "-",
                )
            else:
                log_warn(
                    "Checkpoint missing output file; starting new run.",
                    output_path=output_path or "-",
                )
                start_at = 0
                total_records = None
                output_path = None
                columns = None
        else:
            log_warn(
                "Checkpoint filter differs; starting new run.",
                checkpoint_filter=checkpoint.get("status_filter"),
                current_filter=status_filter,
            )

    length = _normalize_length(length, DEFAULT_RECAP_LENGTH)
    if length > RECAP_LENGTH_MAX:
        log_warn(
            "Recap length above max; clamping.",
            requested=length,
            max=RECAP_LENGTH_MAX,
        )
        length = RECAP_LENGTH_MAX
    if length > RECAP_LENGTH_WARN_THRESHOLD:
        log_warn(
            "Recap page size above recommended; may be capped.",
            length=length,
            fallback=RECAP_LENGTH_WARN_THRESHOLD,
        )

    start_ts = checkpoint.get("run_started_at") if checkpoint else _now_readable()
    if not output_path:
        output_path = _build_output_path(output_dir)

    backup_every_label = backup_every if backup_every > 0 else "off"
    log_info(
        "Rekap",
        status="Running",
        output_path=output_path,
        checkpoint_path=checkpoint_path,
        backup_every=backup_every_label,
    )

    base_payload = _collect_filter_payload(page, status_filter)
    csrf_token = _extract_csrf_token(page)
    if csrf_token:
        base_payload["_token"] = csrf_token
    base_payload["status_filter"] = status_filter

    url = "https://matchapro.web.bps.go.id/direktori-usaha/data-gc-card"
    headers = {
        "Accept": "application/json, text/plain, */*",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": page.url,
    }

    current_start = start_at
    effective_records = (
        effective_records_hint
        if isinstance(effective_records_hint, int) and effective_records_hint > 0
        else total_records
    )
    processed_total = 0
    batch_index = 0
    while True:
        monitor.mark_activity("recap")
        requested_length = length
        payload = dict(base_payload)
        payload["start"] = str(current_start)
        payload["length"] = str(requested_length)

        attempt = 0
        data = None
        records_filtered = None
        status_code = None
        while attempt < max_retries:
            attempt += 1
            try:
                status_code, response_json = _fetch_page(
                    page, url, payload, headers
                )
            except Exception as exc:
                monitor.mark_activity("recap")
                log_warn(
                    "Request failed; retrying.",
                    error=str(exc),
                    attempt=attempt,
                )
                monitor.wait_for_condition(
                    lambda: False, timeout_s=max(0.1, sleep_ms / 1000)
                )
                continue

            if status_code == 419:
                monitor.mark_activity("recap")
                log_warn("CSRF token expired; reloading page.")
                try:
                    page.reload(wait_until="domcontentloaded")
                except Exception:
                    pass
                csrf_token = _extract_csrf_token(page)
                if csrf_token:
                    base_payload["_token"] = csrf_token
                continue

            if status_code != 200 or not isinstance(response_json, dict):
                monitor.mark_activity("recap")
                log_warn(
                    "Unexpected response; retrying.",
                    status=status_code,
                    attempt=attempt,
                )
                monitor.wait_for_condition(
                    lambda: False, timeout_s=max(0.1, sleep_ms / 1000)
                )
                continue

            data = response_json.get("data") or []
            records_filtered = _to_int(response_json.get("recordsFiltered"))
            total_records = _to_int(response_json.get("recordsTotal"))
            if total_records:
                effective_records = total_records
            if records_filtered is not None and records_filtered > 0:
                effective_records = records_filtered
            elif records_filtered == 0 and data:
                records_filtered = None
            break

        if data is None:
            monitor.mark_activity("recap")
            log_error(
                "Failed to fetch data after retries; stopping.",
                start=current_start,
            )
            break

        if not data:
            monitor.mark_activity("recap")
            log_info("No more data returned; stopping.", start=current_start)
            break

        actual_count = len(data)
        next_length = requested_length
        if (
            actual_count < requested_length
            and records_filtered is not None
            and records_filtered > (current_start + actual_count)
        ):
            if requested_length > RECAP_LENGTH_WARN_THRESHOLD:
                next_length = RECAP_LENGTH_WARN_THRESHOLD
                log_warn(
                    "Server capped page size; falling back to safe length.",
                    requested=requested_length,
                    observed=actual_count,
                    next_length=next_length,
                )
            else:
                next_length = actual_count
                log_warn(
                    "Server capped page size; adjusting length.",
                    new_length=next_length,
                )

        captured_at = _now_readable()
        rows = [
            _normalize_record(item, current_start, captured_at) for item in data
        ]
        if columns is None:
            columns = list(rows[0].keys()) if rows else []
        next_batch_index = batch_index + 1
        do_backup = (
            backup_every > 0
            and next_batch_index % backup_every == 0
        )
        _append_xlsx_split(
            output_path,
            rows,
            columns,
            backup_existing=do_backup,
        )
        monitor.mark_activity("recap")
        batch_index += 1
        processed_total += len(rows)
        total_value = (
            effective_records
            if isinstance(effective_records, int) and effective_records > 0
            else None
        )
        log_info(
            "Batch saved.",
            batch=batch_index,
            start=current_start,
            count=len(rows),
            progress=(
                f"{processed_total}/{total_value}"
                if total_value
                else f"{processed_total}/-"
            ),
        )
        if progress_callback:
            total_value = total_value if total_value is not None else 0
            try:
                progress_callback(processed_total, total_value, 0)
            except Exception:
                pass

        length = _normalize_length(next_length, requested_length)
        checkpoint_data = {
            "status": "running",
            "run_started_at": start_ts,
            "last_start": current_start,
            "length": length,
            "backup_every": backup_every,
            "records_total": total_records,
            "records_filtered": records_filtered,
            "status_filter": status_filter,
            "output_path": output_path,
            "columns": columns,
            "split_sheets": list(SPLIT_SHEETS),
            "updated_at": captured_at,
        }
        _save_checkpoint(checkpoint_path, checkpoint_data)

        current_start += actual_count
        if records_filtered is not None and current_start >= records_filtered:
            break
        if records_filtered is None and actual_count < requested_length:
            break

        if sleep_ms:
            monitor.mark_activity("recap")
            monitor.wait_for_condition(
                lambda: False, timeout_s=max(0.1, sleep_ms / 1000)
            )

    if not os.path.exists(output_path):
        raise RuntimeError("Rekap Excel not found; no data saved.")

    checkpoint_done = {
        "status": "done",
        "run_started_at": start_ts,
        "run_finished_at": _now_readable(),
        "length": length,
        "backup_every": backup_every,
        "records_total": total_records,
        "records_filtered": effective_records,
        "status_filter": status_filter,
        "output_path": output_path,
        "columns": columns,
        "split_sheets": list(SPLIT_SHEETS),
    }
    _save_checkpoint(checkpoint_path, checkpoint_done)
    duration = _format_eta_id(time.time() - start_ts_epoch)
    log_info(
        "Rekap selesai",
        status="Done",
        total=processed_total,
        duration=duration,
        output_path=output_path,
    )
    return output_path
